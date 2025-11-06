"""
excel_reverse_engineering/residual_extractor.py
엑셀에서 잔존율 테이블 추출 및 JSON 변환
"""

import openpyxl
from typing import Dict, Optional


class ResidualRateExtractor:
    """엑셀에서 잔존율 테이블 추출"""

    def __init__(self, excel_path: str):
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)

    def extract_all_vehicles(self) -> Dict:
        """
        모든 시트에서 차량별 잔존율 추출

        Returns:
            Dict: {vehicle_id: {24: {10000: 0.65, ...}, ...}, ...}
        """
        all_vehicles = {}

        for sheet in self.workbook.worksheets:
            vehicle_id = self._normalize_vehicle_id(sheet.title)

            print(f"처리 중: {sheet.title} → {vehicle_id}")

            residual_table = self._extract_from_sheet(sheet)

            if residual_table:
                all_vehicles[vehicle_id] = residual_table
                print(f"  ✓ 추출 완료: {len(residual_table)}개 기간")
            else:
                print(f"  ✗ 잔존율 테이블 없음")

        return all_vehicles

    def _normalize_vehicle_id(self, sheet_name: str) -> str:
        """시트 이름을 차량 ID로 변환"""
        # 예: "기아 K5 2.0 프레스티지" → "K5_2024_2.0_PRESTIGE"
        return sheet_name.replace(" ", "_").upper()

    def _extract_from_sheet(self, sheet) -> Optional[Dict]:
        """
        시트에서 잔존율 테이블 추출

        두 가지 패턴 시도:
        1. 계약기간(열) × 주행거리(행)
        2. 주행거리(열) × 계약기간(행)
        """
        # 패턴 1 시도
        table = self._find_pattern_months_cols(sheet)
        if table:
            return table

        # 패턴 2 시도
        table = self._find_pattern_mileage_cols(sheet)
        if table:
            return table

        return None

    def _find_pattern_months_cols(self, sheet) -> Optional[Dict]:
        """
        패턴 1: 계약기간이 열 헤더, 주행거리가 행 헤더

            |  24개월 | 36개월 | 48개월 | 60개월
        ----+--------+--------+--------+--------
        10k |  0.65  |  0.53  |  0.43  |  0.33
        15k |  0.63  |  0.51  |  0.41  |  0.31
        20k |  0.62  |  0.50  |  0.40  |  0.30
        30k |  0.59  |  0.47  |  0.37  |  0.27
        """
        period_patterns = [24, 36, 48, 60]
        mileage_patterns = [10000, 15000, 20000, 30000]

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            row_values = [cell for cell in row if isinstance(cell, (int, float))]

            # 이 행에 24, 36, 48, 60이 모두 있는지
            if all(period in row_values for period in period_patterns):
                # 각 기간의 열 인덱스 찾기
                period_cols = {}
                for col_idx, cell in enumerate(row, 1):
                    if cell in period_patterns:
                        period_cols[cell] = col_idx

                # 다음 행들에서 주행거리별 잔존율 읽기
                residual_data = {}

                for next_row_idx in range(row_idx + 1, row_idx + 10):
                    if next_row_idx > sheet.max_row:
                        break

                    next_row = list(sheet.iter_rows(
                        min_row=next_row_idx,
                        max_row=next_row_idx,
                        values_only=True
                    ))[0]

                    # 첫 열이 주행거리인지 확인
                    first_cell = next_row[0]
                    if isinstance(first_cell, (int, float)) and first_cell in mileage_patterns:
                        mileage = int(first_cell)

                        for period, col_idx in period_cols.items():
                            rate = next_row[col_idx - 1]
                            if isinstance(rate, (int, float)) and 0 < rate <= 1:
                                if period not in residual_data:
                                    residual_data[period] = {}
                                residual_data[period][mileage] = float(rate)

                if residual_data:
                    return residual_data

        return None

    def _find_pattern_mileage_cols(self, sheet) -> Optional[Dict]:
        """
        패턴 2: 주행거리가 열 헤더, 계약기간이 행 헤더

            | 10k  | 15k  | 20k  | 30k
        ----+------+------+------+------
        24  | 0.65 | 0.63 | 0.62 | 0.59
        36  | 0.53 | 0.51 | 0.50 | 0.47
        48  | 0.43 | 0.41 | 0.40 | 0.37
        60  | 0.33 | 0.31 | 0.30 | 0.27
        """
        period_patterns = [24, 36, 48, 60]
        mileage_patterns = [10000, 15000, 20000, 30000]

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            row_values = [cell for cell in row if isinstance(cell, (int, float))]

            # 이 행에 주행거리가 모두 있는지
            if all(m in row_values for m in mileage_patterns):
                # 각 주행거리의 열 인덱스
                mileage_cols = {}
                for col_idx, cell in enumerate(row, 1):
                    if cell in mileage_patterns:
                        mileage_cols[int(cell)] = col_idx

                # 다음 행들에서 계약기간별 잔존율
                residual_data = {}

                for next_row_idx in range(row_idx + 1, row_idx + 10):
                    if next_row_idx > sheet.max_row:
                        break

                    next_row = list(sheet.iter_rows(
                        min_row=next_row_idx,
                        max_row=next_row_idx,
                        values_only=True
                    ))[0]

                    first_cell = next_row[0]
                    if isinstance(first_cell, (int, float)) and first_cell in period_patterns:
                        period = int(first_cell)
                        residual_data[period] = {}

                        for mileage, col_idx in mileage_cols.items():
                            rate = next_row[col_idx - 1]
                            if isinstance(rate, (int, float)) and 0 < rate <= 1:
                                residual_data[period][mileage] = float(rate)

                if residual_data:
                    return residual_data

        return None
