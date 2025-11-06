"""
excel_reverse_engineering/meritz_extractor.py
메리츠캐피탈 엑셀 전용 추출기
"""

import openpyxl
from typing import Dict, Optional, Tuple
import json
from pathlib import Path


class MeritzResidualExtractor:
    """메리츠캐피탈 엑셀에서 잔존율 데이터 추출"""

    def __init__(self, excel_path: str):
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        self.vehicle_sheet = None
        self.residual_sheet = None

        # 시트 찾기
        for sheet in self.workbook.worksheets:
            if '차종' in sheet.title:
                self.vehicle_sheet = sheet
            elif '잔가' in sheet.title:
                self.residual_sheet = sheet

        if not self.vehicle_sheet:
            raise ValueError("'차종' 시트를 찾을 수 없습니다")
        if not self.residual_sheet:
            raise ValueError("'잔가' 시트를 찾을 수 없습니다")

    def extract_all_vehicles(self) -> Tuple[Dict, Dict]:
        """
        모든 차량 데이터 및 잔존율 추출

        Returns:
            Tuple[Dict, Dict]: (vehicle_master, residual_rates)
        """
        print("="*80)
        print("메리츠캐피탈 데이터 추출 시작")
        print("="*80)

        # 1. 잔가 테이블 추출 (등급별)
        print("\n[1/3] 잔가 테이블 추출 중...")
        residual_tables = self._extract_residual_tables()
        print(f"  ✓ {len(residual_tables)}개 캐피탈사 테이블 추출 완료")

        # 2. 주행거리 조정값 추출
        print("\n[2/3] 주행거리 조정값 추출 중...")
        mileage_adjustments = self._extract_mileage_adjustments()
        print(f"  ✓ 주행거리 조정값: {mileage_adjustments}")

        # 3. 차량 정보 및 잔가 등급 추출
        print("\n[3/3] 차량 데이터 추출 중...")
        vehicle_master, residual_rates = self._extract_vehicles_with_residuals(
            residual_tables,
            mileage_adjustments
        )
        print(f"  ✓ {len(vehicle_master)}대 차량 처리 완료")

        return vehicle_master, residual_rates

    def _extract_residual_tables(self) -> Dict[str, Dict]:
        """
        잔가 시트에서 캐피탈사별 잔가 테이블 추출

        Returns:
            Dict: {grade_name: {period: {grade_letter: rate, ...}, ...}, ...}
        """
        tables = {}

        # West 테이블 (Row 48-53, Col 2-12)
        tables['west'] = self._parse_residual_table(
            start_row=49, end_row=54,
            grade_row=48, grade_col_start=2, grade_col_end=13
        )

        # AJ 테이블 (Row 57-62, Col 2-22)
        tables['aj'] = self._parse_residual_table(
            start_row=58, end_row=63,
            grade_row=57, grade_col_start=2, grade_col_end=23
        )

        # APS 테이블 (Row 65-70, Col 2-23)
        tables['aps'] = self._parse_residual_table(
            start_row=66, end_row=71,
            grade_row=65, grade_col_start=2, grade_col_end=24
        )

        # VGS 테이블 (Row 73-77, Col 2-11)
        tables['vgs'] = self._parse_residual_table(
            start_row=74, end_row=78,
            grade_row=73, grade_col_start=2, grade_col_end=12
        )

        return tables

    def _parse_residual_table(self, start_row: int, end_row: int,
                             grade_row: int, grade_col_start: int,
                             grade_col_end: int) -> Dict:
        """
        특정 영역의 잔가 테이블 파싱

        Returns:
            Dict: {period: {grade: rate, ...}, ...}
        """
        table = {}

        # 등급 헤더 읽기 (grade_row)
        grade_header_cells = list(self.residual_sheet.iter_rows(
            min_row=grade_row,
            max_row=grade_row,
            min_col=grade_col_start,
            max_col=grade_col_end,
            values_only=True
        ))[0]

        # 첫 컬럼(기간) 제외하고 등급만 추출
        grades = []
        for cell in grade_header_cells[1:]:  # 첫 컬럼 스킵
            if cell and isinstance(cell, str):
                grades.append(cell)

        # 기간별 데이터 읽기
        for row_idx in range(start_row, end_row):
            row_data = list(self.residual_sheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                min_col=grade_col_start,
                max_col=grade_col_end,
                values_only=True
            ))[0]

            period = row_data[0]  # 첫 컬럼은 기간 (12, 24, 36, 48, 60, 72)

            if not isinstance(period, (int, float)):
                continue

            period = int(period)
            if period not in [24, 36, 48, 60]:  # 12, 72는 제외
                continue

            table[period] = {}

            # 등급별 잔존율 읽기 (첫 컬럼 제외)
            for grade_idx, grade in enumerate(grades):
                rate_value = row_data[grade_idx + 1]  # +1은 첫 컬럼(기간) 스킵
                if isinstance(rate_value, (int, float)) and 0 < rate_value <= 1:
                    table[period][grade] = float(rate_value)

        return table

    def _extract_mileage_adjustments(self) -> Dict[int, float]:
        """
        주행거리별 잔가 조정값 추출

        Returns:
            Dict: {mileage: adjustment, ...}
        """
        adjustments = {}

        # Row 36-39, 컬럼 9-10 (km, W(West) 조정값)
        for row_idx in range(36, 40):
            row_data = list(self.residual_sheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                min_col=9,
                max_col=10,
                values_only=True
            ))[0]

            mileage = row_data[0]  # 컬럼 9: km
            adjustment = row_data[1]  # 컬럼 10: W (West 기준 조정값)

            if isinstance(mileage, (int, float)) and mileage > 0:
                adj_value = float(adjustment) if isinstance(adjustment, (int, float)) else 0.0
                adjustments[int(mileage)] = adj_value

        return adjustments

    def _extract_vehicles_with_residuals(self, residual_tables: Dict,
                                        mileage_adjustments: Dict) -> Tuple[Dict, Dict]:
        """
        차량 정보 및 잔존율 계산

        Returns:
            Tuple[Dict, Dict]: (vehicle_master, residual_rates)
        """
        vehicle_master = {}
        residual_rates = {}

        # 데이터 행 (Row 7부터)
        for row_idx in range(7, self.vehicle_sheet.max_row + 1):
            row_data = list(self.vehicle_sheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                values_only=True
            ))[0]

            # 기본 정보 추출
            maker = row_data[1] if len(row_data) > 1 else None  # Maker (컬럼 B = 2)
            model1 = row_data[2] if len(row_data) > 2 else None  # Model1 (컬럼 C = 3)
            model3 = row_data[4] if len(row_data) > 4 else None  # Model3 (컬럼 E = 5)
            price = row_data[5] if len(row_data) > 5 else None  # 차량가격 (컬럼 F = 6)
            engine_cc = row_data[6] if len(row_data) > 6 else None  # 배기량 (컬럼 G = 7)
            fuel_type = row_data[7] if len(row_data) > 7 else None  # 유종 (컬럼 H = 8)
            west_grade = row_data[9] if len(row_data) > 9 else None  # 웨스트 (컬럼 J = 10)
            aj_grade = row_data[10] if len(row_data) > 10 else None  # 오토준 (컬럼 K = 11)
            aps_grade = row_data[11] if len(row_data) > 11 else None  # APS (컬럼 L = 12)
            vgs_grade = row_data[12] if len(row_data) > 12 else None  # VGS (컬럼 M = 13)

            # 필수 데이터 검증
            if not all([maker, model3, price]):
                continue

            # 차량 ID 생성
            vehicle_id = self._normalize_vehicle_id(maker, model1, model3)

            # 차량 마스터 데이터
            vehicle_master[vehicle_id] = {
                "brand": str(maker),
                "model": str(model1) if model1 else str(maker),
                "trim": str(model3),
                "display_name": f"{maker} {model3}",
                "price": int(price) if isinstance(price, (int, float)) else 0,
                "engine_cc": int(engine_cc) if isinstance(engine_cc, (int, float)) else 0,
                "fuel_type": str(fuel_type) if fuel_type else "unknown",
                "is_import": maker not in ['현대', '기아', '제네시스', 'KGM', '쌍용'],
                "west_grade": str(west_grade) if west_grade else None,
                "aj_grade": str(aj_grade) if aj_grade else None,
                "aps_grade": str(aps_grade) if aps_grade else None,
                "vgs_grade": str(vgs_grade) if vgs_grade else None
            }

            # 모든 등급별 잔존율 계산
            residual_data = {}

            # West 등급
            if west_grade:
                west_normal = self._calculate_residual_for_vehicle(
                    str(west_grade), residual_tables.get('west', {}), mileage_adjustments
                )
                if west_normal:
                    residual_data['west_normal'] = west_normal
                    # 고잔가 = 일반잔가 + 8%
                    residual_data['west_premium'] = self._apply_premium_adjustment(west_normal, 0.08)

            # APS 등급
            if aps_grade:
                aps_normal = self._calculate_residual_for_vehicle(
                    str(aps_grade), residual_tables.get('aps', {}), mileage_adjustments
                )
                if aps_normal:
                    residual_data['aps_normal'] = aps_normal
                    # 고잔가 = 일반잔가 + 8%
                    residual_data['aps_premium'] = self._apply_premium_adjustment(aps_normal, 0.08)

            # VGS 등급
            if vgs_grade:
                vgs_normal = self._calculate_residual_for_vehicle(
                    str(vgs_grade), residual_tables.get('vgs', {}), mileage_adjustments
                )
                if vgs_normal:
                    residual_data['vgs_normal'] = vgs_normal
                    # 고잔가 = 일반잔가 + 6%
                    residual_data['vgs_premium'] = self._apply_premium_adjustment(vgs_normal, 0.06)

            if residual_data:
                residual_rates[vehicle_id] = residual_data

        return vehicle_master, residual_rates

    def _calculate_residual_for_vehicle(self, grade: str, residual_table: Dict,
                                       mileage_adjustments: Dict) -> Optional[Dict]:
        """
        특정 차량의 잔존율 계산

        Returns:
            Dict: {24: {10000: 0.65, 15000: 0.63, ...}, ...}
        """
        result = {}

        for period in [24, 36, 48, 60]:
            if period not in residual_table:
                continue

            if grade not in residual_table[period]:
                continue

            base_rate = residual_table[period][grade]
            result[period] = {}

            # 주행거리별 조정
            for mileage in [10000, 15000, 20000, 30000]:
                adjustment = mileage_adjustments.get(mileage, 0.0)
                adjusted_rate = base_rate + adjustment
                result[period][mileage] = round(max(0.1, min(0.95, adjusted_rate)), 4)

        return result if result else None

    def _apply_premium_adjustment(self, normal_data: Dict, premium_rate: float) -> Dict:
        """
        일반잔가에 고잔가 보정 적용

        Args:
            normal_data: 일반잔가 데이터
            premium_rate: 보정율 (0.08 = +8%p, 0.06 = +6%p)

        Returns:
            Dict: 고잔가 데이터
        """
        premium_data = {}

        for period, mileages in normal_data.items():
            premium_data[period] = {}
            for mileage, rate in mileages.items():
                premium_data[period][mileage] = round(min(0.95, rate + premium_rate), 4)

        return premium_data

    def _normalize_vehicle_id(self, maker: str, model: str, trim: str) -> str:
        """차량 ID 생성"""
        # 공백 제거 및 언더스코어로 변환
        parts = [maker, model if model else '', trim]
        vehicle_id = '_'.join([str(p).strip().replace(' ', '_') for p in parts if p])
        return vehicle_id.upper()


def main():
    """메인 실행 함수"""
    import sys

    if len(sys.argv) < 2:
        print("사용법: python meritz_extractor.py <엑셀파일경로>")
        sys.exit(1)

    excel_path = sys.argv[1]

    # 추출
    extractor = MeritzResidualExtractor(excel_path)
    vehicle_master, residual_rates = extractor.extract_all_vehicles()

    # 결과 저장
    output_dir = Path("data")
    output_dir.mkdir(exist_ok=True)
    (output_dir / "residual_rates").mkdir(exist_ok=True)

    # vehicle_master.json
    vehicle_master_path = output_dir / "vehicle_master.json"
    with open(vehicle_master_path, 'w', encoding='utf-8') as f:
        json.dump(vehicle_master, f, indent=2, ensure_ascii=False)
    print(f"\n✓ 저장: {vehicle_master_path} ({len(vehicle_master)}대)")

    # residual_rates/meritz_capital.json
    # JSON 키를 문자열로 변환
    residual_json = {}
    for vehicle_id, periods in residual_rates.items():
        residual_json[vehicle_id] = {
            str(period): {
                str(mileage): rate
                for mileage, rate in mileages.items()
            }
            for period, mileages in periods.items()
        }

    residual_rates_path = output_dir / "residual_rates" / "meritz_capital.json"
    with open(residual_rates_path, 'w', encoding='utf-8') as f:
        json.dump(residual_json, f, indent=2, ensure_ascii=False)
    print(f"✓ 저장: {residual_rates_path} ({len(residual_rates)}대)")

    # 통계
    print("\n" + "="*80)
    print("추출 완료!")
    print("="*80)
    print(f"총 차량 수: {len(vehicle_master)}대")
    print(f"잔존율 데이터: {len(residual_rates)}대")
    complete_vehicles = sum(1 for v in residual_rates.values() if len(v) == 4)
    print(f"완전한 데이터 (4개 기간): {complete_vehicles}대")

    # 샘플 출력
    if residual_rates:
        print("\n샘플 데이터 (첫 3대):")
        for idx, (vehicle_id, periods) in enumerate(list(residual_rates.items())[:3], 1):
            vehicle = vehicle_master[vehicle_id]
            print(f"\n  [{idx}] {vehicle['display_name']}")
            print(f"      가격: {vehicle['price']:,}원")
            print(f"      등급: {vehicle['west_grade']}")
            print(f"      잔존율: {list(periods.keys())}개월")
            if 36 in periods and 20000 in periods[36]:
                print(f"      36개월/20,000km: {periods[36][20000]:.2%}")


if __name__ == "__main__":
    main()
