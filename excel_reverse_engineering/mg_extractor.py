"""
excel_reverse_engineering/mg_extractor.py
MG캐피탈 엑셀 전용 추출기

메리츠와 완전히 독립적인 데이터 구조
- 잔가map 시트: 등급별 테이블
- 차량DB: 차량별 등급 정보 (S열)
- PMT 계산 방식
"""

import openpyxl
from typing import Dict, Optional, Tuple, List
import json
from pathlib import Path


class MGCapitalExtractor:
    """MG캐피탈 엑셀에서 데이터 추출"""

    def __init__(self, excel_path: str):
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        self.vehicle_db_sheet = None
        self.residual_map_sheet = None

        # 시트 찾기
        for sheet in self.workbook.worksheets:
            if sheet.title == '차량DB':
                self.vehicle_db_sheet = sheet
            elif sheet.title == '잔가map':
                self.residual_map_sheet = sheet

        if not self.vehicle_db_sheet:
            raise ValueError("'차량DB' 시트를 찾을 수 없습니다")
        if not self.residual_map_sheet:
            raise ValueError("'잔가map' 시트를 찾을 수 없습니다")

    def extract_all_data(self) -> Tuple[Dict, Dict]:
        """
        모든 차량 데이터 및 잔존율 추출

        Returns:
            Tuple[Dict, Dict]: (vehicle_master, residual_rates)
        """
        print("="*80)
        print("MG캐피탈 데이터 추출 시작")
        print("="*80)

        # 1. 잔가 등급 테이블 추출 (잔가map 시트)
        print("\n[1/3] 잔가 등급 테이블 추출 중...")
        grade_tables = self._extract_grade_tables()
        print(f"  ✓ {len(grade_tables)}개 등급 테이블 추출 완료")
        for table_name, data in grade_tables.items():
            print(f"    - {table_name}: {len(data['grades'])}개 등급")

        # 2. 차량 마스터 추출
        print("\n[2/3] 차량 마스터 추출 중...")
        vehicle_master = self._extract_vehicle_master()
        print(f"  ✓ {len(vehicle_master)}대 차량 추출 완료")

        # 3. 차량별 잔존율 계산
        print("\n[3/3] 차량별 잔존율 계산 중...")
        residual_rates = self._calculate_vehicle_residuals(
            vehicle_master,
            grade_tables
        )

        # 통계
        vehicles_with_rates = sum(1 for v in residual_rates.values() if v)
        print(f"  ✓ {vehicles_with_rates}/{len(vehicle_master)}대 차량에 잔존율 계산 완료")

        return vehicle_master, residual_rates

    def _extract_grade_tables(self) -> Dict[str, Dict]:
        """
        잔가map 시트에서 등급별 잔가율 테이블 추출

        Returns:
            Dict: {
                'snk': {'grades': [...], 'residual_rates': {...}},
                'aps': {'grades': [...], 'residual_rates': {...}}
            }
        """
        tables = {}

        # 에스앤케이모터스 테이블 (Row 2-7)
        tables['snk'] = self._parse_grade_table(
            name_row=1, name_col=1,  # A1: ■ 에스앤케이모터스
            grade_row=2, grade_col_start=3, grade_col_end=30,  # C2:AD2
            data_start_row=3, data_end_row=7,  # 12/24/36/48/60개월
            period_col=2  # B열에 기간
        )

        # APS 테이블 (Row 13-18)
        tables['aps'] = self._parse_grade_table(
            name_row=12, name_col=1,  # A12: ■ APS
            grade_row=13, grade_col_start=3, grade_col_end=30,  # C13:AD13
            data_start_row=14, data_end_row=18,
            period_col=2
        )

        return tables

    def _parse_grade_table(self, name_row: int, name_col: int,
                           grade_row: int, grade_col_start: int, grade_col_end: int,
                           data_start_row: int, data_end_row: int,
                           period_col: int) -> Dict:
        """
        잔가map에서 하나의 등급 테이블 파싱

        Returns:
            Dict: {
                'name': '에스앤케이모터스',
                'grades': ['A', 'B', 'C', ...],
                'residual_rates': {
                    12: [0.85, 0.84, ...],
                    24: [0.77, 0.76, ...],
                    ...
                }
            }
        """
        ws = self.residual_map_sheet

        # 테이블 이름
        table_name = ws.cell(name_row, name_col).value

        # 등급 추출 (헤더 행)
        grades = []
        for col in range(grade_col_start, grade_col_end + 1):
            grade = ws.cell(grade_row, col).value
            if grade and grade not in [None, '']:
                grades.append(str(grade))
            else:
                break  # 빈 셀 만나면 중단

        # 잔가율 데이터 추출 (기간별)
        residual_rates = {}
        for row in range(data_start_row, data_end_row + 1):
            period = ws.cell(row, period_col).value
            if not period or not isinstance(period, (int, float)):
                continue

            period = int(period)
            rates = []

            for col_idx, grade in enumerate(grades, start=grade_col_start):
                rate = ws.cell(row, col_idx).value
                if rate is not None and isinstance(rate, (int, float)):
                    rates.append(round(float(rate), 4))
                else:
                    rates.append(None)

            residual_rates[period] = rates

        return {
            'name': table_name,
            'grades': grades,
            'residual_rates': residual_rates
        }

    def _extract_vehicle_master(self) -> Dict:
        """
        차량DB에서 차량 마스터 데이터 추출

        Returns:
            Dict: {
                vehicle_id: {
                    'brand': 'BMW',
                    'model': 'X5 30d...',
                    'displacement': 2993,
                    'vehicle_type': '승용SUV(7~10인)',
                    'price': 115500000,
                    'grade_snk': 'E',  # S열 (col 19)
                    'premium_available': 'Y'  # P열 (col 16)
                }
            }
        """
        ws = self.vehicle_db_sheet
        vehicles = {}

        # 헤더 행 찾기 (BRAND, MODEL 포함된 행)
        header_row = None
        for row_idx in range(1, 10):
            row_values = [ws.cell(row_idx, col).value for col in range(1, 20)]
            if 'BRAND' in row_values and 'MODEL' in row_values:
                header_row = row_idx
                break

        if not header_row:
            raise ValueError("차량DB 헤더를 찾을 수 없습니다")

        # 데이터 추출 (헤더 다음 행부터)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            brand = ws.cell(row_idx, 5).value  # E열 (col 5)
            model = ws.cell(row_idx, 6).value  # F열 (col 6)

            if not brand or not model:
                continue  # 빈 행 건너뛰기

            # 차량 ID 생성 (BRAND_MODEL, 공백/특수문자 제거)
            vehicle_id = self._generate_vehicle_id(brand, model)

            # model을 기본 모델과 트림으로 분리 (간단히 첫 단어를 model로)
            model_parts = str(model).split(' ', 1)
            base_model = model_parts[0] if len(model_parts) > 0 else str(model)
            trim = model_parts[1] if len(model_parts) > 1 else ""

            # 배기량 확인
            displacement = ws.cell(row_idx, 7).value
            engine_cc = int(displacement) if displacement and isinstance(displacement, (int, float)) else 0

            # 연료 타입 추정 (전기차는 배기량 0)
            if engine_cc == 0:
                fuel_type = "전기"
            elif "hybrid" in str(model).lower() or "plug" in str(model).lower():
                fuel_type = "하이브리드"
            else:
                fuel_type = "휘발유/경유"

            # 수입차 여부 (국산 브랜드 목록)
            domestic_brands = ["현대", "기아", "제네시스", "쌍용", "르노코리아", "KGM"]
            is_import = brand not in domestic_brands

            # 차량 정보 (메리츠 호환 구조)
            vehicles[vehicle_id] = {
                'brand': str(brand),
                'model': base_model,
                'trim': trim,
                'display_name': f"{brand} {model}",
                'displacement': displacement,  # G열: 배기량
                'engine_cc': engine_cc,
                'fuel_type': fuel_type,
                'vehicle_type': ws.cell(row_idx, 8).value,  # H열: 차종분류
                'price': ws.cell(row_idx, 9).value,  # I열: 차량가
                'is_import': is_import,
                'grade_snk': ws.cell(row_idx, 19).value,  # S열: 에스앤케이모터스 등급
                'premium_available': ws.cell(row_idx, 16).value  # P열: 고잔가 가능
            }

        return vehicles

    def _generate_vehicle_id(self, brand: str, model: str) -> str:
        """차량 ID 생성 (BRAND_MODEL 형식)"""
        # 공백, 특수문자 제거 및 언더스코어로 변환
        vehicle_id = f"{brand}_{model}"
        vehicle_id = vehicle_id.replace(' ', '_')
        vehicle_id = vehicle_id.replace('(', '').replace(')', '')
        vehicle_id = vehicle_id.replace('/', '_')
        vehicle_id = vehicle_id.replace('-', '_')
        vehicle_id = vehicle_id.upper()
        return vehicle_id

    def _calculate_vehicle_residuals(self, vehicle_master: Dict,
                                     grade_tables: Dict) -> Dict:
        """
        차량별 잔존율 계산

        MG 특징:
        1. 등급 테이블에서 기본 잔가율 조회
        2. 주행거리 보정 적용
        3. 고잔가 보정 (+8%) 적용

        Returns:
            Dict: {
                vehicle_id: {
                    'snk_normal': {24: {10000: 0.77, 20000: 0.77, ...}, ...},
                    'snk_premium': {24: {10000: 0.85, 20000: 0.85, ...}, ...}
                }
            }
        """
        residual_rates = {}
        snk_table = grade_tables['snk']
        periods = [12, 24, 36, 48, 60]
        mileages = [10000, 15000, 20000, 30000]

        for vehicle_id, vehicle_data in vehicle_master.items():
            grade_snk = vehicle_data.get('grade_snk')
            premium_available = vehicle_data.get('premium_available')

            if not grade_snk or grade_snk not in snk_table['grades']:
                # 등급 정보 없으면 건너뛰기
                residual_rates[vehicle_id] = None
                continue

            # 등급 인덱스
            grade_idx = snk_table['grades'].index(grade_snk)

            # SNK 일반 잔가
            snk_normal = {}
            for period in periods:
                if period not in snk_table['residual_rates']:
                    continue

                base_rate = snk_table['residual_rates'][period][grade_idx]
                if base_rate is None:
                    continue

                snk_normal[period] = {}
                for mileage in mileages:
                    # 주행거리 보정 (SNK 기준)
                    adjusted_rate = self._apply_mileage_adjustment_snk(
                        base_rate, mileage
                    )
                    snk_normal[period][mileage] = round(adjusted_rate, 4)

            # SNK 고잔가 (+8%)
            snk_premium = None
            if premium_available == 'Y':
                snk_premium = {}
                for period, mileage_dict in snk_normal.items():
                    snk_premium[period] = {}
                    for mileage, rate in mileage_dict.items():
                        # 고잔가 = 일반잔가 + 8%
                        premium_rate = round(min(0.95, rate + 0.08), 4)
                        snk_premium[period][mileage] = premium_rate

            residual_rates[vehicle_id] = {
                'snk_normal': snk_normal,
                'snk_premium': snk_premium
            }

        return residual_rates

    def _apply_mileage_adjustment_snk(self, base_rate: float, mileage: int) -> float:
        """
        SNK 주행거리 보정 (CQ59 공식 참조)

        =IF($BD$26=35000, $CP$59-2%,
            IF($BD$26=30000, $CP$59-4%,
                IF($BD$26=20000, $CP$59,
                    IF($BD$26=10000, $CP$59+0.5%, 0%)
                )
            )
        )
        """
        if mileage == 10000:
            return base_rate + 0.005  # +0.5%
        elif mileage == 15000:
            # 15,000은 공식에 없음, 20,000과 동일하게 처리
            return base_rate
        elif mileage == 20000:
            return base_rate  # 기준
        elif mileage == 30000:
            return base_rate - 0.04  # -4%
        elif mileage == 35000:
            return base_rate - 0.02  # -2%
        else:
            return base_rate


def main():
    """메인 실행 함수"""
    import sys

    if len(sys.argv) < 2:
        print("사용법: python mg_extractor.py <엑셀파일경로>")
        print("예시: python mg_extractor.py xlsx/mg_capital_2510_vol3.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_dir = Path("data")

    try:
        # 추출
        extractor = MGCapitalExtractor(excel_path)
        vehicle_master, residual_rates = extractor.extract_all_data()

        # 저장
        print("\n" + "="*80)
        print("데이터 저장 중...")
        print("="*80)

        # 차량 마스터 저장
        vehicle_output = output_dir / "mg_vehicle_master.json"
        with open(vehicle_output, 'w', encoding='utf-8') as f:
            json.dump(vehicle_master, f, ensure_ascii=False, indent=2)
        print(f"✓ 차량 마스터: {vehicle_output} ({len(vehicle_master)}대)")

        # 잔존율 저장
        residual_output = output_dir / "residual_rates" / "mg_capital.json"
        residual_output.parent.mkdir(exist_ok=True)
        with open(residual_output, 'w', encoding='utf-8') as f:
            json.dump(residual_rates, f, ensure_ascii=False, indent=2)

        vehicles_with_rates = sum(1 for v in residual_rates.values() if v)
        print(f"✓ 잔존율: {residual_output} ({vehicles_with_rates}대)")

        print("\n" + "="*80)
        print("✅ 추출 완료!")
        print("="*80)

    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
