"""
tools/analyze_residual_sheet.py
잔가 시트 상세 분석 도구
"""

import openpyxl
import sys


def analyze_residual_sheet(excel_path: str):
    """잔가 시트 상세 분석"""
    print("="*80)
    print(f"잔가 시트 분석: {excel_path}")
    print("="*80)

    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    # 잔가 시트 찾기
    residual_sheet = None
    for sheet in workbook.worksheets:
        if '잔가' in sheet.title:
            residual_sheet = sheet
            break

    if not residual_sheet:
        print("❌ '잔가' 시트를 찾을 수 없습니다.")
        return

    print(f"\n잔가 시트 발견: {residual_sheet.title}")
    print(f"크기: {residual_sheet.max_row} 행 × {residual_sheet.max_column} 열")

    # 전체 데이터 출력
    print("\n전체 데이터:")
    for row_idx in range(1, min(residual_sheet.max_row + 1, 80)):
        row = list(residual_sheet.iter_rows(
            min_row=row_idx,
            max_row=row_idx,
            values_only=True
        ))[0]
        # None이 아닌 값만 표시
        non_empty = []
        for idx, cell in enumerate(row[:23], 1):
            if cell is not None and str(cell).strip():
                non_empty.append(f"[{idx}]{str(cell)[:50]}")

        if non_empty:
            print(f"Row {row_idx:2d}: {' | '.join(non_empty)}")

    # 숫자 패턴 검색
    print("\n\n숫자 패턴 검색:")
    period_patterns = [24, 36, 48, 60]
    mileage_patterns = [10000, 15000, 20000, 30000]

    for row_idx, row in enumerate(residual_sheet.iter_rows(values_only=True), 1):
        row_values = [cell for cell in row if isinstance(cell, (int, float))]

        # 계약기간 패턴
        if all(period in row_values for period in period_patterns):
            print(f"✓ 계약기간 패턴 발견 (24, 36, 48, 60) - Row {row_idx}")
            print(f"  전체 값: {row[:23]}")

        # 주행거리 패턴
        if all(m in row_values for m in mileage_patterns):
            print(f"✓ 주행거리 패턴 발견 (10000, 15000, 20000, 30000) - Row {row_idx}")
            print(f"  전체 값: {row[:23]}")


def analyze_vehicle_sheet(excel_path: str):
    """차종 시트 상세 분석"""
    print("\n" + "="*80)
    print(f"차종 시트 분석")
    print("="*80)

    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    # 차종 시트 찾기
    vehicle_sheet = None
    for sheet in workbook.worksheets:
        if '차종' in sheet.title:
            vehicle_sheet = sheet
            break

    if not vehicle_sheet:
        print("❌ '차종' 시트를 찾을 수 없습니다.")
        return

    print(f"\n차종 시트 발견: {vehicle_sheet.title}")
    print(f"크기: {vehicle_sheet.max_row} 행 × {vehicle_sheet.max_column} 열")

    # 헤더 분석 (Row 6)
    header_row = list(vehicle_sheet.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    print("\n헤더 (Row 6, 첫 30개 컬럼):")
    for idx, header in enumerate(header_row[:30], 1):
        if header:
            print(f"  [{idx}] {header}")

    # 샘플 데이터 (Row 7-12)
    print("\n샘플 차량 데이터 (Row 7-12, 첫 15개 컬럼):")
    for row_idx in range(7, 13):
        row = list(vehicle_sheet.iter_rows(
            min_row=row_idx,
            max_row=row_idx,
            values_only=True
        ))[0]
        print(f"Row {row_idx}: {row[:15]}")

    # 잔존율 데이터 위치 찾기
    print("\n\n잔존율 데이터 위치 탐색:")
    print("(24, 36, 48, 60 패턴 또는 0.3~0.7 범위의 연속된 소수 찾기)")

    # 헤더에서 기간 관련 컬럼 찾기
    for idx, header in enumerate(header_row[:100], 1):
        if header and isinstance(header, (int, float)):
            if header in [24, 36, 48, 60]:
                print(f"  ✓ 컬럼 [{idx}]: {header} (계약기간으로 추정)")
        elif header and isinstance(header, str):
            if any(keyword in str(header) for keyword in ['24', '36', '48', '60', '잔가', '잔존', 'RV']):
                print(f"  ✓ 컬럼 [{idx}]: {header}")

    # 데이터 행에서 잔존율 패턴 찾기 (0.3~0.7 범위의 소수)
    sample_row = list(vehicle_sheet.iter_rows(min_row=7, max_row=7, values_only=True))[0]
    potential_rv_cols = []
    for idx, value in enumerate(sample_row[:100], 1):
        if isinstance(value, (int, float)) and 0.2 <= value <= 0.9:
            potential_rv_cols.append((idx, value))

    if potential_rv_cols:
        print(f"\n잔존율로 추정되는 컬럼들 (0.2~0.9 범위):")
        for col_idx, value in potential_rv_cols[:20]:
            print(f"  컬럼 [{col_idx}]: {value}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python analyze_residual_sheet.py <엑셀파일경로>")
        sys.exit(1)

    analyze_residual_sheet(sys.argv[1])
    analyze_vehicle_sheet(sys.argv[1])
