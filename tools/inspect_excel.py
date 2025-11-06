"""
tools/inspect_excel.py
엑셀 파일 구조 확인 도구
"""

import openpyxl
import sys


def inspect_excel(excel_path: str):
    """엑셀 파일 구조 확인"""
    print("="*80)
    print(f"엑셀 파일 분석: {excel_path}")
    print("="*80)

    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    print(f"\n총 시트 수: {len(workbook.worksheets)}")
    print("\n시트 목록:")

    for idx, sheet in enumerate(workbook.worksheets, 1):
        print(f"\n[{idx}] 시트 이름: {sheet.title}")
        print(f"    크기: {sheet.max_row} 행 × {sheet.max_column} 열")

        # 첫 10행 미리보기
        print("    첫 10행 데이터 미리보기:")
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            # None이 아닌 값만 필터링
            non_empty = [str(cell)[:30] if cell is not None else '' for cell in row[:10]]
            if any(non_empty):
                print(f"    Row {row_idx:2d}: {non_empty}")

        # 숫자 패턴 찾기 (24, 36, 48, 60)
        period_patterns = [24, 36, 48, 60]
        mileage_patterns = [10000, 15000, 20000, 30000]

        print("\n    패턴 검색:")
        period_found = False
        mileage_found = False

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            row_values = [cell for cell in row if isinstance(cell, (int, float))]

            # 계약기간 패턴
            if not period_found and all(period in row_values for period in period_patterns):
                print(f"    ✓ 계약기간 패턴 발견 (24, 36, 48, 60) - Row {row_idx}")
                period_found = True

            # 주행거리 패턴
            if not mileage_found and all(m in row_values for m in mileage_patterns):
                print(f"    ✓ 주행거리 패턴 발견 (10000, 15000, 20000, 30000) - Row {row_idx}")
                mileage_found = True

            if period_found and mileage_found:
                break

        if not period_found:
            print("    ✗ 계약기간 패턴 없음")
        if not mileage_found:
            print("    ✗ 주행거리 패턴 없음")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python inspect_excel.py <엑셀파일경로>")
        sys.exit(1)

    inspect_excel(sys.argv[1])
