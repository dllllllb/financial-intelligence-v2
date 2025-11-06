"""
tools/excel_validator.py
엑셀 견적서와 계산 결과 비교 검증
"""

import openpyxl
import sys
from pathlib import Path

# 상위 디렉토리를 path에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.calculator import calculate_operating_lease, calculate_auto_tax
from data import vehicle_master, residual_rates, interest_rates


def extract_lease_quote_from_excel(excel_path: str, sheet_name: str = "운용리스 입력") -> dict:
    """
    엑셀 견적서에서 조건 및 결과 추출

    Args:
        excel_path: 엑셀 파일 경로
        sheet_name: 시트 이름

    Returns:
        dict: 견적 조건 및 결과
    """
    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'를 찾을 수 없습니다")

    sheet = workbook[sheet_name]

    # 메리츠캐피탈 "운용리스 입력" 시트 구조 기반
    # Row 2: 일자, 등록명의
    # Row 4-5: 기간, 주행거리
    # Row 7-14: 차량 정보, 조건
    # Row 19: 리스료 결과

    result = {}

    try:
        # 브랜드 (Row 5, Col D)
        result['brand'] = sheet.cell(5, 4).value

        # 세부 모델 (Row 7, Col D)
        result['model'] = sheet.cell(7, 4).value

        # 차량가 (Row 11, Col D)
        result['vehicle_price'] = sheet.cell(11, 4).value

        # 계약 기간 (Row 4, Col J) - 조건1
        result['contract_months'] = sheet.cell(4, 10).value

        # 주행거리 (Row 5, Col J)
        result['annual_mileage'] = sheet.cell(5, 10).value

        # 보증금 (Row 7, Col J)
        result['down_payment'] = sheet.cell(7, 10).value or 0

        # 잔가 (Row 11, Col J)
        result['residual_rate'] = sheet.cell(11, 10).value

        # 엑셀 계산 결과 - 리스료 (Row 19, Col J)
        # 위치는 엑셀 구조에 따라 조정 필요
        result['excel_monthly_lease'] = None

        # 여러 위치 시도
        for row in range(15, 25):
            cell_value = sheet.cell(row, 10).value
            if isinstance(cell_value, (int, float)) and 100_000 < cell_value < 10_000_000:
                result['excel_monthly_lease'] = cell_value
                break

        # 추가: 운용리스 내부 시트에서 더 정확한 데이터 추출
        if '운용리스 내부' in workbook.sheetnames:
            internal_sheet = workbook['운용리스 내부']

            # 취득원가 (Row 22, Col D in 내부 sheet)
            acquisition_cost = internal_sheet.cell(22, 4).value
            if acquisition_cost:
                result['acquisition_cost'] = acquisition_cost

            # 금리 (Row 32, Col H in 내부 sheet - Col G는 라벨)
            interest_rate = internal_sheet.cell(32, 8).value
            if interest_rate and isinstance(interest_rate, (int, float)):
                result['interest_rate'] = interest_rate

            # 리스료 (Row 27, Col H in 내부 sheet - 더 정확함)
            monthly_lease = internal_sheet.cell(27, 8).value
            if monthly_lease and isinstance(monthly_lease, (int, float)):
                result['excel_monthly_lease'] = monthly_lease

    except Exception as e:
        print(f"⚠ 데이터 추출 중 오류: {e}")

    return result


def compare_with_excel(
    excel_path: str,
    capital_id: str = "meritz_capital",
    tolerance: float = 2000
) -> dict:
    """
    엑셀 견적서와 계산 결과 비교

    Args:
        excel_path: 엑셀 파일 경로
        capital_id: 캐피탈 ID
        tolerance: 허용 오차 (원)

    Returns:
        dict: 비교 결과
    """
    print("="*80)
    print("엑셀 견적서 검증")
    print("="*80)

    # 엑셀에서 데이터 추출
    print(f"\n[1] 엑셀 데이터 추출: {excel_path}")
    excel_data = extract_lease_quote_from_excel(excel_path)

    print(f"  브랜드: {excel_data.get('brand')}")
    print(f"  모델: {excel_data.get('model')}")
    print(f"  차량가: {excel_data.get('vehicle_price'):,}원")
    if excel_data.get('acquisition_cost'):
        print(f"  취득원가: {excel_data.get('acquisition_cost'):,}원")
    print(f"  계약기간: {excel_data.get('contract_months')}개월")
    print(f"  주행거리: {excel_data.get('annual_mileage'):,}km")
    print(f"  잔존율: {excel_data.get('residual_rate'):.2%}")
    if excel_data.get('interest_rate'):
        rate = float(excel_data.get('interest_rate'))
        print(f"  엑셀 금리: {rate:.4%}")

    if excel_data.get('excel_monthly_lease'):
        print(f"  엑셀 리스료: {excel_data['excel_monthly_lease']:,}원")
    else:
        print(f"  ⚠ 엑셀 리스료를 찾을 수 없습니다")

    # 차량 검색
    print(f"\n[2] 차량 데이터 조회")
    vehicles = vehicle_master.search_vehicles(excel_data['model'])

    if not vehicles:
        print(f"  ✗ 차량을 찾을 수 없습니다: {excel_data['model']}")
        return {"success": False, "error": "차량 없음"}

    vehicle_id = vehicles[0]['id']
    vehicle = vehicle_master.get_vehicle(vehicle_id)
    print(f"  ✓ 차량 ID: {vehicle_id}")
    print(f"  ✓ DB 가격: {vehicle['price']:,}원")
    print(f"  ✓ 엑셀 가격: {excel_data['vehicle_price']:,}원 (계산에 사용)")

    # 금리 조회
    print(f"\n[3] 금리 조회")
    if excel_data.get('interest_rate'):
        # 엑셀에서 추출한 금리 사용
        annual_rate = excel_data['interest_rate']
        print(f"  ✓ 엑셀 금리 사용: {annual_rate:.4%}")
    else:
        # DB에서 금리 조회
        annual_rate = interest_rates.get_interest_rate(
            capital_id=capital_id,
            vehicle_price=excel_data['vehicle_price'],
            brand=vehicle['brand'],
            is_import=vehicle['is_import'],
            is_ev=(vehicle['engine_cc'] == 0),
            contract_months=excel_data['contract_months']
        )
        print(f"  ✓ DB 금리 사용: {annual_rate:.2%}")

    # 자동차세 계산
    annual_car_tax = calculate_auto_tax(vehicle['engine_cc'], True)
    print(f"  ✓ 연간 자동차세: {annual_car_tax:,.0f}원")

    # 리스료 계산
    print(f"\n[4] 리스료 계산")

    # 취득원가가 있으면 하이브리드 방식 사용
    if excel_data.get('acquisition_cost'):
        print(f"  방식: 하이브리드 (잔존가치=차량가 기준, 금융=취득원가 기준)")
        result = calculate_operating_lease(
            vehicle_price=excel_data['vehicle_price'],
            contract_months=excel_data['contract_months'],
            down_payment=excel_data['down_payment'],
            residual_rate=excel_data['residual_rate'],
            annual_rate=annual_rate,
            acquisition_tax_rate=0.0,
            registration_fee=0,  # 이미 취득원가에 포함
            annual_car_tax=annual_car_tax,
            method='simple',
            acquisition_cost=excel_data['acquisition_cost']
        )
    else:
        print(f"  방식: 표준 (차량가 기준)")
        result = calculate_operating_lease(
            vehicle_price=excel_data['vehicle_price'],
            contract_months=excel_data['contract_months'],
            down_payment=excel_data['down_payment'],
            residual_rate=excel_data['residual_rate'],
            annual_rate=annual_rate,
            acquisition_tax_rate=0.0,
            registration_fee=200_000,
            annual_car_tax=annual_car_tax,
            method='simple'
        )

    print(f"  ✓ 계산 완료: {result['monthly_total']:,}원")

    # 비교
    print(f"\n[5] 결과 비교")
    print("="*80)

    if excel_data.get('excel_monthly_lease'):
        diff = abs(result['monthly_total'] - excel_data['excel_monthly_lease'])
        diff_percentage = (diff / excel_data['excel_monthly_lease']) * 100

        print(f"\n  엑셀 견적서:  {excel_data['excel_monthly_lease']:>12,}원")
        print(f"  계산 결과:    {result['monthly_total']:>12,}원")
        print(f"  차이:         {diff:>12,}원 ({diff_percentage:+.2f}%)")

        if diff <= tolerance:
            print(f"\n  ✅ 검증 성공! (오차 ±{tolerance:,}원 이내)")
            success = True
        else:
            print(f"\n  ❌ 검증 실패 (오차 ±{tolerance:,}원 초과)")
            success = False

    else:
        print(f"\n  ⚠ 엑셀 리스료를 찾을 수 없어 비교 불가")
        success = False

    print("\n" + "="*80)

    return {
        "success": success,
        "excel_value": excel_data.get('excel_monthly_lease'),
        "calculated_value": result['monthly_total'],
        "difference": diff if excel_data.get('excel_monthly_lease') else None,
        "tolerance": tolerance,
        "details": result
    }


def main():
    """메인 실행"""
    if len(sys.argv) < 2:
        print("사용법: python excel_validator.py <엑셀파일경로> [허용오차]")
        print("예시: python excel_validator.py meritz.xlsx 2000")
        sys.exit(1)

    excel_path = sys.argv[1]
    tolerance = float(sys.argv[2]) if len(sys.argv) > 2 else 2000

    result = compare_with_excel(excel_path, tolerance=tolerance)

    if result['success']:
        print("\n✅ 검증 성공!")
        sys.exit(0)
    else:
        print("\n❌ 검증 실패")
        sys.exit(1)


if __name__ == "__main__":
    main()
