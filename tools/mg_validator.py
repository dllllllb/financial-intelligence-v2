"""
tools/mg_validator.py
MG 엑셀 견적서 검증 도구

MG 엑셀과 우리 계산기 결과를 비교
"""

import openpyxl
from core.mg_calculator import MGLeaseCalculator
import sys


def validate_mg_excel(excel_path: str, tolerance: int = 2000):
    """
    MG 엑셀 견적서 검증

    Args:
        excel_path: MG 엑셀 파일 경로
        tolerance: 허용 오차 (원)

    Returns:
        bool: 검증 성공 여부
    """
    print("=" * 100)
    print("MG캐피탈 엑셀 검증")
    print("=" * 100)

    # 엑셀 로드
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["운용리스"]

    # 1. 차량 정보 추출
    print("\n[1] 엑셀에서 데이터 추출 중...")
    print("-" * 100)

    vehicle_name = ws.cell(8, 4).value  # D8
    vehicle_price_str = ws.cell(14, 11).value  # K14 (차량가격)

    # 문자열에서 숫자 추출
    if isinstance(vehicle_price_str, str):
        vehicle_price = int(vehicle_price_str.replace(',', '').replace('원', ''))
    else:
        vehicle_price = int(vehicle_price_str)

    print(f"차량: {vehicle_name}")
    print(f"차량가: {vehicle_price:,}원")

    # 2. 계약 조건 추출
    contract_months = int(ws.cell(9, 33).value)  # AG9
    annual_mileage = int(ws.cell(15, 33).value)  # AG15
    residual_rate = float(ws.cell(16, 33).value)  # AG16 (잔존율)

    print(f"\n계약 조건:")
    print(f"  계약 기간: {contract_months}개월")
    print(f"  주행거리: {annual_mileage:,}km/년")
    print(f"  잔존율: {residual_rate:.4f} ({residual_rate*100:.2f}%)")

    # 3. 금리 추출 (BD37)
    annual_interest_rate = float(ws.cell(37, 56).value)  # BD37 (col 56)
    print(f"  금리: {annual_interest_rate:.6f} ({annual_interest_rate*100:.4f}%)")

    # 4. 선납금 추출 (AG17, AG18, AG11)
    # MG는 보증금/장기선수금/선납리스료로 나뉨
    # 현재 엑셀에서는 모두 0
    down_payment_guarantee = ws.cell(17, 33).value or 0  # AG17: 보증금
    down_payment_advance = ws.cell(18, 33).value or 0    # AG18: 장기선수금
    down_payment_lease = ws.cell(11, 33).value or 0      # AG11: 선납리스료
    total_down_payment = int(down_payment_guarantee + down_payment_advance + down_payment_lease)

    print(f"  선납금: {total_down_payment:,}원")

    # 5. 엑셀 계산 결과
    excel_monthly_payment = ws.cell(14, 33).value  # AG14 (月납입금)
    if isinstance(excel_monthly_payment, str):
        excel_monthly_payment = int(excel_monthly_payment.replace(',', '').replace('원', ''))
    else:
        excel_monthly_payment = int(excel_monthly_payment)

    print(f"\n엑셀 계산 결과:")
    print(f"  월 납입금: {excel_monthly_payment:,}원")

    # 6. 우리 계산기로 계산
    print("\n" + "-" * 100)
    print("[2] 우리 계산기로 계산 중...")
    print("-" * 100)

    calc = MGLeaseCalculator()

    # 취득원가 추출 (검증용)
    acquisition_cost_excel = ws.cell(19, 11).value  # K19
    if isinstance(acquisition_cost_excel, str):
        acquisition_cost_excel = int(acquisition_cost_excel.replace(',', '').replace('원', ''))
    else:
        acquisition_cost_excel = int(acquisition_cost_excel)

    # 선납금 비율 계산
    down_payment_rate = total_down_payment / acquisition_cost_excel if total_down_payment > 0 else 0.0

    result = calc.calculate(
        vehicle_price=vehicle_price,
        residual_rate=residual_rate,
        contract_months=contract_months,
        annual_mileage=annual_mileage,
        annual_interest_rate=annual_interest_rate,
        down_payment_rate=down_payment_rate,
        region="서울",
        is_ev=False,  # TODO: 전기차 판별 로직
        is_hybrid=False
    )

    print(f"우리 계산 결과:")
    print(f"  월 납입금: {result['monthly_payment']:,}원")
    print(f"  취득원가: {result['acquisition_cost']:,}원")
    print(f"  잔존가치: {result['residual_value']:,}원")

    # 7. 비교
    print("\n" + "=" * 100)
    print("[3] 검증 결과")
    print("=" * 100)

    diff = abs(result['monthly_payment'] - excel_monthly_payment)
    diff_pct = (diff / excel_monthly_payment) * 100 if excel_monthly_payment > 0 else 0

    print(f"\n월 납입금 비교:")
    print(f"  엑셀:     {excel_monthly_payment:,}원")
    print(f"  계산기:   {result['monthly_payment']:,}원")
    print(f"  차이:     {diff:,}원 ({diff_pct:.3f}%)")
    print(f"  허용오차: ±{tolerance:,}원")

    # 취득원가 비교 (참고)
    acquisition_diff = abs(result['acquisition_cost'] - acquisition_cost_excel)
    print(f"\n취득원가 비교 (참고):")
    print(f"  엑셀:     {acquisition_cost_excel:,}원")
    print(f"  계산기:   {result['acquisition_cost']:,}원")
    print(f"  차이:     {acquisition_diff:,}원")

    # 잔존가치 비교 (참고)
    residual_value_excel = int(ws.cell(16, 36).value)  # AJ16
    residual_diff = abs(result['residual_value'] - residual_value_excel)
    print(f"\n잔존가치 비교 (참고):")
    print(f"  엑셀:     {residual_value_excel:,}원")
    print(f"  계산기:   {result['residual_value']:,}원")
    print(f"  차이:     {residual_diff:,}원")

    # 판정
    print("\n" + "=" * 100)
    if diff <= tolerance:
        print(f"✅ 검증 성공! (오차 {diff:,}원 ≤ {tolerance:,}원)")
        print("=" * 100)
        return True
    else:
        print(f"❌ 검증 실패! (오차 {diff:,}원 > {tolerance:,}원)")
        print("=" * 100)
        return False


def main():
    if len(sys.argv) < 2:
        print("사용법: python mg_validator.py <엑셀파일경로> [허용오차]")
        print("예시: python mg_validator.py xlsx/mg_capital_2510_vol3.xlsx 2000")
        sys.exit(1)

    excel_path = sys.argv[1]
    tolerance = int(sys.argv[2]) if len(sys.argv) > 2 else 2000

    try:
        success = validate_mg_excel(excel_path, tolerance)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
