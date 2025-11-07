"""
MG Capital - 등급 컬럼(S열) 확인
"""
import openpyxl
import sys

def find_grade_column(filepath: str):
    print("=" * 100)
    print("차량DB - S열 (col 19) 확인: 등급 정보!")
    print("=" * 100)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["차량DB"]

    # 헤더 확인 (Row 1-5)
    print("\nRow 1-5 (헤더):")
    for row in range(1, 6):
        val = ws.cell(row, 19).value  # S열 = col 19
        if val:
            print(f"  S{row}: {val}")

    # 실제 데이터 확인 (Row 6-25, 20개 샘플)
    print("\n차량 데이터 (Row 6-25, 샘플):")
    print(f"{'Row':<5} {'F(MODEL)':<40} {'S(등급)':<10}")
    print("-" * 100)

    for row in range(6, 26):
        model = ws.cell(row, 6).value  # F열
        grade = ws.cell(row, 19).value  # S열

        model_str = str(model)[:38] if model else ""
        grade_str = str(grade) if grade else "(없음)"

        print(f"{row:<5} {model_str:<40} {grade_str:<10}")

    # BMW X5 찾기
    print("\nBMW X5 등급 확인:")
    for row in range(6, 700):
        model = ws.cell(row, 6).value
        if model and "X5" in str(model) and "30d" in str(model):
            brand = ws.cell(row, 5).value
            grade = ws.cell(row, 19).value
            price = ws.cell(row, 9).value
            print(f"  Row {row}: {brand} {model}")
            if price:
                print(f"    가격: {price:,}원")
            else:
                print("    가격: None")
            print(f"    등급(S열): {grade}")
            break

    # 전체 등급 종류 확인
    print("\n전체 등급 종류 (S열, 중복 제거):")
    grades = set()
    for row in range(6, 700):
        grade = ws.cell(row, 19).value
        if grade and grade not in ["", None]:
            grades.add(str(grade))

    grades_sorted = sorted(list(grades))
    print(f"  총 {len(grades_sorted)}개 등급: {grades_sorted}")

    print("\n=" * 100)

if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else "xlsx/mg_capital_2510_vol3.xlsx"
    find_grade_column(filepath)
