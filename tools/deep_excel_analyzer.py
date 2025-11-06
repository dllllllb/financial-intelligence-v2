"""
tools/deep_excel_analyzer.py
금융사 엑셀 파일 심층 분석 도구
"""

import openpyxl
import sys
from typing import Dict, List, Any


class ExcelDeepAnalyzer:
    """엑셀 파일 구조 심층 분석"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)

    def analyze(self):
        """전체 분석 실행"""
        print("="*100)
        print(f"엑셀 파일 심층 분석: {self.excel_path}")
        print("="*100)

        print(f"\n총 시트 수: {len(self.workbook.worksheets)}")

        for idx, sheet in enumerate(self.workbook.worksheets, 1):
            print("\n" + "="*100)
            print(f"[시트 {idx}/{len(self.workbook.worksheets)}] {sheet.title}")
            print("="*100)

            self._analyze_sheet(sheet)

    def _analyze_sheet(self, sheet):
        """개별 시트 분석"""
        print(f"\n📊 기본 정보:")
        print(f"   - 크기: {sheet.max_row} 행 × {sheet.max_column} 열")

        # 1. 데이터 밀도 분석
        print(f"\n📈 데이터 밀도 분석:")
        total_cells = sheet.max_row * min(sheet.max_column, 50)  # 첫 50개 컬럼만
        non_empty = 0

        for row in sheet.iter_rows(max_col=50, values_only=True):
            non_empty += sum(1 for cell in row if cell is not None and str(cell).strip())

        density = (non_empty / total_cells * 100) if total_cells > 0 else 0
        print(f"   - 데이터 밀도: {density:.1f}% (비어있지 않은 셀)")

        # 2. 숫자 패턴 검색
        print(f"\n🔍 중요 패턴 검색:")
        self._find_patterns(sheet)

        # 3. 헤더 행 추정
        print(f"\n📋 헤더 추정:")
        self._find_headers(sheet)

        # 4. 테이블 구조 추정
        print(f"\n📐 테이블 구조:")
        self._find_tables(sheet)

        # 5. 샘플 데이터 (첫 20행, 중요 컬럼만)
        print(f"\n📄 샘플 데이터 (첫 20행):")
        self._show_sample_data(sheet, max_rows=20)

    def _find_patterns(self, sheet):
        """중요 패턴 검색"""
        period_patterns = [24, 36, 48, 60]
        mileage_patterns = [10000, 15000, 20000, 30000]

        found_periods = []
        found_mileages = []
        found_residuals = []  # 잔존율 범위 (0.2~0.9)

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True, max_row=100), 1):
            row_values = [cell for cell in row[:50] if isinstance(cell, (int, float))]

            # 계약기간 패턴
            if all(period in row_values for period in period_patterns):
                found_periods.append(row_idx)

            # 주행거리 패턴
            if all(m in row_values for m in mileage_patterns):
                found_mileages.append(row_idx)

            # 잔존율 패턴 (연속된 0.2~0.9 값)
            residual_values = [v for v in row_values if 0.2 <= v <= 0.9]
            if len(residual_values) >= 4:
                found_residuals.append((row_idx, residual_values[:10]))

        if found_periods:
            print(f"   ✓ 계약기간 패턴 (24,36,48,60): Row {found_periods}")

        if found_mileages:
            print(f"   ✓ 주행거리 패턴 (10k,15k,20k,30k): Row {found_mileages}")

        if found_residuals:
            print(f"   ✓ 잔존율 패턴 (0.2~0.9 연속값):")
            for row_idx, values in found_residuals[:3]:
                print(f"      Row {row_idx}: {[f'{v:.2f}' for v in values]}")

    def _find_headers(self, sheet):
        """헤더 행 찾기"""
        potential_headers = []

        for row_idx in range(1, min(20, sheet.max_row + 1)):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # 문자열이 많고, 일정 수 이상의 비어있지 않은 셀이 있는 행
            text_cells = [cell for cell in row[:30] if isinstance(cell, str) and len(str(cell).strip()) > 0]

            if len(text_cells) >= 3:  # 최소 3개 이상의 텍스트 셀
                potential_headers.append({
                    'row': row_idx,
                    'count': len(text_cells),
                    'sample': text_cells[:5]
                })

        if potential_headers:
            print(f"   추정 헤더 행:")
            for header in potential_headers[:3]:
                print(f"      Row {header['row']} ({header['count']}개 컬럼): {header['sample']}")

    def _find_tables(self, sheet):
        """테이블 구조 찾기"""
        # 연속된 데이터 블록 찾기
        data_blocks = []
        current_block = None

        for row_idx in range(1, min(sheet.max_row + 1, 100)):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            non_empty = sum(1 for cell in row[:30] if cell is not None and str(cell).strip())

            if non_empty >= 3:  # 데이터가 있는 행
                if current_block is None:
                    current_block = {'start': row_idx, 'end': row_idx}
                else:
                    current_block['end'] = row_idx
            else:  # 빈 행
                if current_block and (current_block['end'] - current_block['start']) >= 3:
                    data_blocks.append(current_block)
                current_block = None

        if current_block and (current_block['end'] - current_block['start']) >= 3:
            data_blocks.append(current_block)

        if data_blocks:
            print(f"   발견된 데이터 블록 (3행 이상):")
            for block in data_blocks[:5]:
                rows = block['end'] - block['start'] + 1
                print(f"      Row {block['start']}~{block['end']} ({rows}행)")

    def _show_sample_data(self, sheet, max_rows=20):
        """샘플 데이터 출력"""
        for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # 의미있는 데이터만 표시
            display_data = []
            for col_idx, cell in enumerate(row[:30], 1):  # 첫 30개 컬럼만
                if cell is not None and str(cell).strip():
                    value_str = str(cell)
                    if len(value_str) > 40:
                        value_str = value_str[:40] + "..."
                    display_data.append(f"[{col_idx}]{value_str}")

            if display_data:
                print(f"   Row {row_idx:3d}: {' | '.join(display_data)}")


def main():
    """메인 실행"""
    if len(sys.argv) < 2:
        print("사용법: python deep_excel_analyzer.py <엑셀파일경로>")
        sys.exit(1)

    analyzer = ExcelDeepAnalyzer(sys.argv[1])
    analyzer.analyze()


if __name__ == "__main__":
    main()
